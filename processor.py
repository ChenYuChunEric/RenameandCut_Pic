import os
import csv
import re
from PIL import Image
import openpyxl
import xlrd

class PhotoProcessor:
    @staticmethod
    def _read_csv_rows(csv_path, max_rows=None):
        """
        嘗試不同編碼讀取 CSV 檔案。
        """
        encodings = ['utf-8-sig', 'utf-8', 'cp950']
        for enc in encodings:
            try:
                with open(csv_path, 'r', encoding=enc, newline='') as f:
                    reader = csv.reader(f)
                    rows = []
                    for i, row in enumerate(reader):
                        if max_rows is not None and i >= max_rows:
                            break
                        rows.append(row)
                    return rows
            except (UnicodeDecodeError, LookupError):
                continue
        raise RuntimeError("無法識別 CSV 檔案的編碼格式，請確認是否為 UTF-8 或 Big5 (CP950) 格式。")

    @staticmethod
    def get_excel_headers(excel_path):
        """
        讀取 Excel/CSV 檔案的首行，獲取所有欄位名稱。
        """
        if not excel_path or not os.path.exists(excel_path):
            raise FileNotFoundError("找不到指定的對照檔案。")
        
        _, ext = os.path.splitext(excel_path.lower())
        
        try:
            if ext == '.csv':
                rows = PhotoProcessor._read_csv_rows(excel_path, max_rows=1)
                if rows:
                    return [str(cell).strip() for cell in rows[0] if cell is not None]
                return []
                
            elif ext == '.xls':
                wb = xlrd.open_workbook(excel_path)
                sheet = wb.sheet_by_index(0)
                if sheet.nrows > 0:
                    return [str(cell.value).strip() for cell in sheet.row(0) if cell.value is not None]
                return []
                
            else:
                wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                sheet = wb.active
                for row in sheet.iter_rows(max_row=1, values_only=True):
                    headers = [str(cell).strip() for cell in row if cell is not None]
                    return headers
                return []
        except Exception as e:
            raise RuntimeError(f"讀取對照檔標頭失敗: {str(e)}")

    @staticmethod
    def load_excel_mapping(excel_path, id_column, student_id_column):
        """
        讀取 Excel/CSV 建立身分證字號到學號的對照字典。
        """
        if not excel_path or not os.path.exists(excel_path):
            raise FileNotFoundError("找不到指定的對照檔案。")
        
        _, ext = os.path.splitext(excel_path.lower())
        mapping = {}
        
        try:
            if ext == '.csv':
                rows = PhotoProcessor._read_csv_rows(excel_path)
                if not rows:
                    return {}
                headers = [str(cell).strip() for cell in rows[0] if cell is not None]
                if id_column not in headers or student_id_column not in headers:
                    raise ValueError("找不到指定的欄位名稱，請確認 CSV 欄位設定。")
                id_idx = headers.index(id_column)
                student_id_idx = headers.index(student_id_column)
                
                for row in rows[1:]:
                    if len(row) > max(id_idx, student_id_idx):
                        id_val = row[id_idx]
                        student_id_val = row[student_id_idx]
                        if id_val is not None and student_id_val is not None:
                            id_str = str(id_val).strip()
                            student_id_str = str(student_id_val).strip()
                            # 去除 Excel 將數字讀為 float 產生的 .0
                            if id_str.endswith('.0'):
                                id_str = id_str[:-2]
                            if student_id_str.endswith('.0'):
                                student_id_str = student_id_str[:-2]
                            if id_str:
                                mapping[id_str] = student_id_str
                                
            elif ext == '.xls':
                wb = xlrd.open_workbook(excel_path)
                sheet = wb.sheet_by_index(0)
                if sheet.nrows == 0:
                    return {}
                headers = [str(cell.value).strip() for cell in sheet.row(0) if cell.value is not None]
                if id_column not in headers or student_id_column not in headers:
                    raise ValueError("找不到指定的欄位名稱，請確認 Excel 欄位設定。")
                id_idx = headers.index(id_column)
                student_id_idx = headers.index(student_id_column)
                
                for r in range(1, sheet.nrows):
                    row = sheet.row(r)
                    if len(row) > max(id_idx, student_id_idx):
                        id_val = row[id_idx].value
                        student_id_val = row[student_id_idx].value
                        if id_val is not None and student_id_val is not None:
                            id_str = str(id_val).strip()
                            student_id_str = str(student_id_val).strip()
                            if id_str.endswith('.0'):
                                id_str = id_str[:-2]
                            if student_id_str.endswith('.0'):
                                student_id_str = student_id_str[:-2]
                            if id_str:
                                mapping[id_str] = student_id_str
                                
            else:
                wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                sheet = wb.active
                headers = []
                id_idx = -1
                student_id_idx = -1
                
                for row in sheet.iter_rows(max_row=1, values_only=True):
                    headers = [str(cell).strip() for cell in row if cell is not None]
                    break
                    
                if id_column in headers:
                    id_idx = headers.index(id_column)
                if student_id_column in headers:
                    student_id_idx = headers.index(student_id_column)
                    
                if id_idx == -1 or student_id_idx == -1:
                    raise ValueError("找不到指定的欄位名稱，請確認 Excel 欄位設定。")
                
                is_first = True
                for row in sheet.iter_rows(values_only=True):
                    if is_first:
                        is_first = False
                        continue
                    
                    if len(row) > max(id_idx, student_id_idx):
                        id_val = row[id_idx]
                        student_id_val = row[student_id_idx]
                        
                        if id_val is not None and student_id_val is not None:
                            id_str = str(id_val).strip()
                            student_id_str = str(student_id_val).strip()
                            if id_str.endswith('.0'):
                                id_str = id_str[:-2]
                            if student_id_str.endswith('.0'):
                                student_id_str = student_id_str[:-2]
                            if id_str:
                                mapping[id_str] = student_id_str
                                
            return mapping
        except Exception as e:
            raise RuntimeError(f"讀取對照資料失敗: {str(e)}")

    @staticmethod
    def scan_source_directory(dir_path):
        """
        掃描來源資料夾中的所有圖片檔案。
        """
        if not dir_path or not os.path.exists(dir_path):
            return []
            
        supported_exts = {'.jpg', '.jpeg', '.png', '.bmp', '.webp', '.tiff'}
        photos = []
        
        try:
            for file_name in os.listdir(dir_path):
                full_path = os.path.join(dir_path, file_name)
                if os.path.isfile(full_path):
                    name_part, ext = os.path.splitext(file_name)
                    if ext.lower() in supported_exts:
                        # 擷取第一個符合 A-Za-z 加上 9 個數字 的身分證字號，並轉大寫
                        match = re.search(r'([A-Za-z]\d{9})', name_part)
                        photo_id = match.group(1).upper() if match else name_part.strip()
                        photos.append({
                            'filename': file_name,
                            'filepath': full_path,
                            'id': photo_id,
                            'size': os.path.getsize(full_path)
                        })
            return photos
        except Exception as e:
            raise RuntimeError(f"掃描資料夾失敗: {str(e)}")

    @staticmethod
    def calculate_crop_box(orig_width, orig_height, target_width, target_height, scale=1.0, vertical_offset=0.0):
        """
        計算在原圖上的裁切框座標 (x1, y1, x2, y2)。
        - scale: 縮放比例 (0.1 ~ 1.0)，代表裁切框占最大可用區域的比例。比例越小，最終圖片人像越大（Zoom-in）。
        - vertical_offset: 垂直偏移量 (-0.5 ~ 0.5)，正值代表裁切框向上移動（保留上方），負值向下移動。
        """
        target_ratio = target_width / target_height
        orig_ratio = orig_width / orig_height
        
        # 計算最大可用裁切尺寸 (等比例填滿)
        if orig_ratio < target_ratio:
            # 原圖比目標更窄高，以寬度為基準
            max_crop_w = orig_width
            max_crop_h = orig_width / target_ratio
        else:
            # 原圖比目標更寬，以高度為基準
            max_crop_h = orig_height
            max_crop_w = orig_height * target_ratio
            
        # 依據縮放比例計算實際裁切框寬高
        crop_w = max_crop_w * scale
        crop_h = max_crop_h * scale
        
        # 預設中心點 (原圖中心)
        center_x = orig_width / 2.0
        center_y = orig_height / 2.0
        
        # 套用垂直位移 (限制在安全範圍內，避免裁切框超出原圖)
        # 垂直偏移量乘以原圖高度作為像素偏移
        pixel_offset_y = -vertical_offset * orig_height  # 習慣上：正值往上移，所以減去 offset
        center_y += pixel_offset_y
        
        # 計算初步的左上角與右下角座標
        x1 = center_x - crop_w / 2.0
        y1 = center_y - crop_h / 2.0
        x2 = center_x + crop_w / 2.0
        y2 = center_y + crop_h / 2.0
        
        # 邊界限制 (Clamp) 並平移裁切框，使其不超出原圖
        if x1 < 0:
            x2 -= x1
            x1 = 0.0
        if x2 > orig_width:
            x1 -= (x2 - orig_width)
            x2 = float(orig_width)
            
        if y1 < 0:
            y2 -= y1
            y1 = 0.0
        if y2 > orig_height:
            y1 -= (y2 - orig_height)
            y2 = float(orig_height)
            
        # 再次確保不越界 (防止計算誤差)
        x1 = max(0.0, min(x1, orig_width))
        y1 = max(0.0, min(y1, orig_height))
        x2 = max(x1, min(x2, orig_width))
        y2 = max(y1, min(y2, orig_height))
        
        return int(x1), int(y1), int(x2), int(y2)

    @classmethod
    def process_image(cls, src_path, dest_path, target_width, target_height, scale=1.0, vertical_offset=0.0, output_format='JPEG'):
        """
        處理單張圖片（轉檔、裁切、縮放）。
        - dest_path: 如果為 None，則不存檔，僅回傳 PIL Image 物件（用於預覽）。
        - output_format: 'JPEG' 或 'PNG'。
        """
        if not src_path or not os.path.exists(src_path):
            raise FileNotFoundError("找不到來源相片檔案。")
            
        try:
            with Image.open(src_path) as img:
                # 處理圖片旋轉 (部分手機拍攝的照片會有 EXIF 旋轉標記)
                img = cls._correct_image_orientation(img)
                
                orig_w, orig_h = img.size
                
                # 計算裁切框
                x1, y1, x2, y2 = cls.calculate_crop_box(
                    orig_w, orig_h, target_width, target_height, scale, vertical_offset
                )
                
                # 進行裁切
                cropped_img = img.crop((x1, y1, x2, y2))
                
                # 縮放至目標解析度 (使用高品質重採樣)
                resized_img = cropped_img.resize((target_width, target_height), Image.Resampling.LANCZOS)
                
                # 如果是預覽，直接回傳
                if dest_path is None:
                    return resized_img
                
                # 確保輸出資料夾存在
                os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                
                # 轉換為 RGB 模式 (若要存為 JPEG，且原圖是 RGBA 格式)
                if output_format.upper() in ('JPEG', 'JPG'):
                    if resized_img.mode in ('RGBA', 'LA') or (resized_img.mode == 'P' and 'transparency' in resized_img.info):
                        bg = Image.new('RGB', resized_img.size, (255, 255, 255))
                        bg.paste(resized_img, mask=resized_img.convert('RGBA').split()[3])
                        resized_img = bg
                    else:
                        resized_img = resized_img.convert('RGB')
                    
                    resized_img.save(dest_path, format='JPEG', quality=95)
                else:
                    resized_img.save(dest_path, format='PNG')
                    
                return True
        except Exception as e:
            raise RuntimeError(f"處理圖片失敗 ({os.path.basename(src_path)}): {str(e)}")

    @staticmethod
    def _correct_image_orientation(img):
        """
        根據 EXIF 資訊修正圖片旋轉方向。
        """
        try:
            # EXIF Orientation Tag ID 為 274
            exif = img._getexif()
            if exif is not None and 274 in exif:
                orientation = exif[274]
                if orientation == 3:
                    return img.rotate(180, expand=True)
                elif orientation == 6:
                    return img.rotate(270, expand=True)
                elif orientation == 8:
                    return img.rotate(90, expand=True)
        except Exception:
            # 忽略 EXIF 讀取錯誤，直接回傳原圖
            pass
        return img
